import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from apps.authentication.models import TdcUser
from apps.competitions.models import Trial
from apps.attempts.models import Attempt

from openpyxl.utils import get_column_letter

def export_leaderboard_excel() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Classement Général TDC"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="082F6A", end_color="082F6A", fill_type="solid")
    sub_header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    trials = list(Trial.objects.all().order_by('order', 'id'))
    participants = list(TdcUser.objects.filter(role='PARTICIPANT').order_by('participant_code'))

    # Headers
    headers = ["Rang", "Code Participant", "Nom & Prénom", "Groupe / Équipe"]
    for t in trials:
        headers.append(f"Épreuve {t.order}: {t.title} (/{t.max_score} pts)")
    headers.extend(["Score Total Obtenu", "Score Total Max", "Pourcentage Global (%)", "Épreuves Terminées", "Temps Total (min)"])

    # Title row (merged across all columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="TECOX DIGITAL CHALLENGE (TDC) — CLASSEMENT ET RÉSULTATS GÉNÉRAUX")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.append([]) # empty row
    ws.append(headers)
    ws.row_dimensions[3].height = 25

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col_idx)
        c.font = header_font
        c.fill = sub_header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Compute stats for each participant
    rows_data = []
    total_possible_all_trials = sum(t.max_score for t in trials) or 1.0

    for p in participants:
        attempts_map = {a.trial_id: a for a in Attempt.objects.filter(participant=p, is_final=True)}
        p_total_score = 0.0
        p_total_seconds = 0
        p_completed_count = 0
        trial_scores = []

        for t in trials:
            att = attempts_map.get(t.id)
            if att and att.status in ['submitted', 'graded', 'expired']:
                trial_scores.append(att.total_score)
                p_total_score += att.total_score
                p_total_seconds += att.time_spent_seconds
                p_completed_count += 1
            else:
                trial_scores.append(0.0)

        global_pct = round((p_total_score / total_possible_all_trials) * 100, 2)
        total_minutes = round(p_total_seconds / 60, 1)

        rows_data.append({
            'participant': p,
            'trial_scores': trial_scores,
            'total_score': round(p_total_score, 2),
            'total_possible': total_possible_all_trials,
            'global_percentage': global_pct,
            'completed_count': p_completed_count,
            'total_minutes': total_minutes
        })

    # Sort by total score descending
    rows_data.sort(key=lambda x: (x['total_score'], -x['total_minutes']), reverse=True)

    # Fill table
    for rank, row_data in enumerate(rows_data, start=1):
        p = row_data['participant']
        row_cells = [
            f"#{rank}",
            p.participant_code or '',
            p.full_name,
            p.team_group or '-'
        ]
        row_cells.extend(row_data['trial_scores'])
        row_cells.extend([
            row_data['total_score'],
            row_data['total_possible'],
            f"{row_data['global_percentage']}%",
            f"{row_data['completed_count']} / {len(trials)}",
            row_data['total_minutes']
        ])
        ws.append(row_cells)

        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 20

        # Style data cells
        for col_idx in range(1, len(row_cells) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border_thin
            if col_idx in [1, 2, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Top 3 highlights
            if rank == 1:
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Gold tint
            elif rank == 2:
                cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Silver tint
            elif rank == 3:
                cell.fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Bronze tint

    # Auto adjust column widths safely
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(3, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_leaderboard_csv() -> str:
    trials = list(Trial.objects.all().order_by('order', 'id'))
    participants = list(TdcUser.objects.filter(role='PARTICIPANT').order_by('participant_code'))

    output = io.StringIO()
    writer = csv.writer(output)

    headers = ["Rang", "Code Participant", "Nom & Prénom", "Groupe"]
    for t in trials:
        headers.append(f"Épreuve {t.order}: {t.title} ({t.max_score} pts)")
    headers.extend(["Score Total", "Score Max", "Pourcentage", "Épreuves Terminées", "Temps (min)"])
    writer.writerow(headers)

    total_possible_all_trials = sum(t.max_score for t in trials) or 1.0
    rows_data = []

    for p in participants:
        attempts_map = {a.trial_id: a for a in Attempt.objects.filter(participant=p, is_final=True)}
        p_total_score = 0.0
        p_total_seconds = 0
        p_completed_count = 0
        trial_scores = []

        for t in trials:
            att = attempts_map.get(t.id)
            if att and att.status in ['submitted', 'graded', 'expired']:
                trial_scores.append(att.total_score)
                p_total_score += att.total_score
                p_total_seconds += att.time_spent_seconds
                p_completed_count += 1
            else:
                trial_scores.append(0.0)

        global_pct = round((p_total_score / total_possible_all_trials) * 100, 2)
        total_minutes = round(p_total_seconds / 60, 1)

        rows_data.append({
            'code': p.participant_code or '',
            'name': p.full_name,
            'team': p.team_group or '',
            'trial_scores': trial_scores,
            'total_score': round(p_total_score, 2),
            'total_possible': total_possible_all_trials,
            'global_percentage': f"{global_pct}%",
            'completed_count': f"{p_completed_count}/{len(trials)}",
            'total_minutes': total_minutes
        })

    rows_data.sort(key=lambda x: (x['total_score'], -x['total_minutes']), reverse=True)

    for rank, row in enumerate(rows_data, start=1):
        line = [rank, row['code'], row['name'], row['team']]
        line.extend(row['trial_scores'])
        line.extend([row['total_score'], row['total_possible'], row['global_percentage'], row['completed_count'], row['total_minutes']])
        writer.writerow(line)

    return output.getvalue()
